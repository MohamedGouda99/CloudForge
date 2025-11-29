#!/usr/bin/env python3
"""
Generate TypeScript resource catalogs from Cloud_Services Excel data.

Outputs frontend/src/lib/resources/generatedCatalog.ts with every service
listed in Cloud_Services/Cloud_All_Services_Catalog.xlsx, normalising
categories and assigning reasonable fallback icons per category.
"""

from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
CLOUD_SERVICES_PATH = Path(
    os.environ.get("CLOUD_SERVICES_PATH", PROJECT_ROOT / "Cloud_Services")
)
EXCEL_PATH = CLOUD_SERVICES_PATH / "Cloud_All_Services_Catalog.xlsx"
OUTPUT_PATH = (
    PROJECT_ROOT / "frontend" / "src" / "lib" / "resources" / "generatedCatalog.ts"
)

PROVIDER_SHEETS = {
    "AWS": "aws",
    "Azure": "azure",
    "GCP": "gcp",
}

PRIMARY_CATEGORY_MAP: Dict[str, str] = {
    "account management": "security",
    "ai/ml": "analytics",
    "ai/ml platform": "analytics",
    "analytics": "analytics",
    "api": "application",
    "archive": "storage",
    "audit": "security",
    "backup": "storage",
    "best practices": "application",
    "bi": "analytics",
    "cdn": "networking",
    "compliance": "security",
    "compute": "compute",
    "config": "application",
    "containers": "containers",
    "conversational ai": "analytics",
    "cost management": "application",
    "data exchange": "analytics",
    "data governance": "analytics",
    "data integration": "analytics",
    "data lake": "analytics",
    "data preparation": "analytics",
    "data processing": "analytics",
    "data warehouse": "database",
    "database": "database",
    "dev tools": "application",
    "devops": "application",
    "dns": "networking",
    "edge": "networking",
    "engagement": "application",
    "event bus": "application",
    "governance": "application",
    "hybrid": "networking",
    "ide": "application",
    "identity": "security",
    "iot": "application",
    "maps": "application",
    "media": "application",
    "messaging": "application",
    "metadata": "analytics",
    "migration": "application",
    "mobile": "application",
    "monitoring": "analytics",
    "observability": "analytics",
    "ops": "application",
    "orchestration": "application",
    "paas": "compute",
    "security": "security",
    "serverless": "compute",
    "service mesh": "networking",
    "storage": "storage",
    "tracing": "analytics",
    "zero trust": "security",
}

KEYWORD_CATEGORY_MAP: List[Tuple[str, str]] = [
    ("kubernetes", "containers"),
    ("container", "containers"),
    ("queue", "application"),
    ("pubsub", "application"),
    ("messaging", "application"),
    ("search", "analytics"),
    ("observability", "analytics"),
    ("monitor", "analytics"),
    ("analytics", "analytics"),
    ("govern", "application"),
    ("identity", "security"),
    ("secrets", "security"),
    ("key ", "security"),
    ("iot", "application"),
    ("edge", "networking"),
    ("graph", "database"),
    ("sql", "database"),
    ("database", "database"),
    ("storage", "storage"),
    ("backup", "storage"),
    ("archive", "storage"),
    ("cdn", "networking"),
    ("dns", "networking"),
    ("network", "networking"),
    ("load balancer", "networking"),
    ("lb", "networking"),
    ("security", "security"),
    ("devops", "application"),
    ("dev tools", "application"),
    ("artifact", "application"),
    ("git", "application"),
    ("ci", "application"),
    ("cd", "application"),
    ("ops", "application"),
    ("runbook", "application"),
    ("scheduler", "application"),
    ("app service", "application"),
    ("app ", "application"),
    ("compute", "compute"),
    ("serverless", "compute"),
    ("lambda", "compute"),
    ("function", "compute"),
    ("fargate", "containers"),
]


DEFAULT_CATEGORY_ICONS: Dict[str, str] = {
    "compute": "lucide:cpu",
    "storage": "lucide:hard-drive",
    "database": "lucide:database",
    "networking": "lucide:network",
    "security": "lucide:shield",
    "analytics": "lucide:chart-line",
    "application": "lucide:app-window",
    "containers": "lucide:boxes",
    "other": "lucide:circle",
}


@dataclass
class GeneratedResource:
    type: str
    label: str
    provider: str
    category: str
    original_category: str
    description: str
    icon: str

    def to_dict(self) -> Dict[str, str]:
        return {
            "type": self.type,
            "label": self.label,
            "provider": self.provider,
            "category": self.category,
            "originalCategory": self.original_category,
            "description": self.description,
            "icon": self.icon,
        }


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = value.replace("&", "and")
    value = re.sub(r"[^\w\s-]+", "", value)
    value = re.sub(r"[\s/]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def normalise_category(raw: str | None) -> Tuple[str, str]:
    if not raw:
        return "other", ""

    original = raw.strip()
    raw_lower = original.lower()
    primary = raw_lower.split("/")[0].strip()

    if primary in PRIMARY_CATEGORY_MAP:
        return PRIMARY_CATEGORY_MAP[primary], original

    for keyword, mapped in KEYWORD_CATEGORY_MAP:
        if keyword in raw_lower:
            return mapped, original

    return "other", original


def ensure_unique(base: str, used: Iterable[str]) -> str:
    if base not in used:
        return base
    counter = 2
    while f"{base}_{counter}" in used:
        counter += 1
    return f"{base}_{counter}"


def load_excel_catalog() -> Dict[str, List[GeneratedResource]]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel catalog not found at {EXCEL_PATH}")

    workbook = openpyxl.load_workbook(EXCEL_PATH)
    catalog: Dict[str, List[GeneratedResource]] = defaultdict(list)

    for sheet_name, provider in PROVIDER_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        used_types: set[str] = set()

        for service_name_cell, category_cell, *_ in worksheet.iter_rows(min_row=2):
            service_name = (service_name_cell.value or "").strip()
            if not service_name:
                continue

            category_raw = category_cell.value if category_cell is not None else None
            category, original_category = normalise_category(category_raw)

            base_type = f"{provider}_{slugify(service_name)}"
            unique_type = ensure_unique(base_type, used_types)
            used_types.add(unique_type)

            icon = DEFAULT_CATEGORY_ICONS.get(category, DEFAULT_CATEGORY_ICONS["other"])

            catalog[provider].append(
                GeneratedResource(
                    type=unique_type,
                    label=service_name,
                    provider=provider,
                    category=category,
                    original_category=original_category,
                    description=f"{service_name} service",
                    icon=icon,
                )
            )

        catalog[provider].sort(key=lambda item: item.label.lower())

    return catalog


def write_typescript(catalog: Dict[str, List[GeneratedResource]]) -> None:
    icon_map: Dict[str, str] = {}
    serialisable_catalog = {
        provider: [resource.to_dict() for resource in resources]
        for provider, resources in catalog.items()
    }

    for resources in catalog.values():
        for resource in resources:
            icon_map[resource.type] = resource.icon

    header = (
        "// Auto-generated by scripts/generate_resource_catalog.py. "
        "Do not edit manually.\n"
        "// Source of truth: Cloud_Services/Cloud_All_Services_Catalog.xlsx\n\n"
        "export type ProviderId = 'aws' | 'azure' | 'gcp';\n\n"
        "export interface GeneratedResourceDefinition {\n"
        "  type: string;\n"
        "  label: string;\n"
        "  provider: ProviderId;\n"
        "  category: string;\n"
        "  originalCategory: string;\n"
        "  description: string;\n"
        "  icon: string;\n"
        "}\n\n"
    )

    catalog_ts = json.dumps(
        serialisable_catalog, indent=2, ensure_ascii=False, sort_keys=True
    )
    icon_map_ts = json.dumps(icon_map, indent=2, ensure_ascii=False, sort_keys=True)

    content = (
        header
        + "export const GENERATED_RESOURCE_CATALOG: Record<ProviderId, GeneratedResourceDefinition[]> = "
        + catalog_ts
        + ";\n\n"
        + "export const GENERATED_ICON_MAP: Record<string, string> = "
        + icon_map_ts
        + ";\n"
    )

    OUTPUT_PATH.write_text(content, encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} with:")
    for provider, resources in catalog.items():
        print(f"  {provider}: {len(resources)} services")


def main() -> None:
    catalog = load_excel_catalog()
    write_typescript(catalog)


if __name__ == "__main__":
    main()
